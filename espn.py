from espn_api.football import League
from pathlib import Path
from math import ceil
import argparse
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
import pandas as pd
import pprint
import numpy as np
import copy
import platform
import requests
import re

ap = argparse.ArgumentParser(description='A script to gather information about your FFB league. It writes the spreadsheets to the $HOME\\ff_league\\ directory.\n\
                                          This should be usable by all ESPN leagues. The ONLY hard coded thing you will need to change is your league\'s keeper \
                                          escalation value. That variable is player_esc_value.')
ap._action_groups.pop()
required = ap.add_argument_group('required arguments')
required.add_argument('-l', '--leagueid', type=int, required=True, help='The numerical league ID from your ESPN Fantasy league', )
required.add_argument('-s', '--swid', type=str, required=True, help='The SWID value found in the cookies of your browser when looking at the league')
required.add_argument('-e', '--espns2', type=str, required=True, help='The espn_s2 value found in the cookies of your browser when looking at the league')
args = vars(ap.parse_args())

league_id = args['leagueid'] # your league_id
print('Here is your league id: {}'.format(league_id))
swid = args['swid'] # your swid
espn_s2 = args['espns2'] # your espn_s2
url = 'https://fantasy.espn.com/apis/v3/games/ffl/seasons/'

def init():
    home = str(Path.home())
    if platform.system() == "Windows":
        write_path = f'{home}\\ff_league\\'
    else:
        write_path = f'{home}/ff_league/'
    print (f'OS: {platform.system()} Path: {write_path}')
    os.system(f'mkdir {write_path}')
    answer = '' # Primed for our prompt
    return write_path, answer

# ESPN player API request
def espn_player_request(player_id, year):
    player_filter = "{\"players\": {\"filterIds\": {\"value\": " + f"[{player_id}]" + "}}}"
    headers = {'X-Fantasy-Source': 'kona',
               'X-Fantasy-Filter': player_filter,
               'Accept': 'application/json'
               }
    cookies = {'SWID': f'{swid}', 'espn_s2': f'{espn_s2}'}
    r = requests.get(url + f'{year}/segments/0/leagues/{league_id}', params={'view': 'kona_playercard'}, 
                             headers=headers, cookies=cookies)
    player_json = r.json()

    return player_json

def get_league_info(answer, year):     
    if answer == 'l':
        year = year - 1
    
    league = {}
    league = League(league_id, year, espn_s2, swid)
    league_info = {}
    for d in range(len(league.teams)):
        team_info = {}
        team_info = league.teams[d].__dict__
        team_name = (team_info['team_name'])
        league_info[team_name] = copy.deepcopy(team_info)

    if answer == 'l':
        basic_league_info(league_info)  

    elif answer == 'k':
        league_df = pd.DataFrame.from_dict(league_info)  
        
        cookies = {'SWID': f'{swid}', 'espn_s2': f'{espn_s2}', '_dcf': '1', 'region': 'unknown'}
        params = {'view': {'mSettings','mTeam','modular','mNav'}}
        print('Checking to see if your league uses keepers')
        r = requests.get(url + f'{year}/segments/0/leagues/{league_id}', params=params, cookies=cookies)
        settings = r.json()
        if settings['settings']['draftSettings']['keeperCount'] > 0:
            update = ''
            while update.lower() not in ('y', 'n'):
                update = input('You are using keepers, would you like to update your keeper values on ESPN? [Y/N] ')
                if update.lower() in ('y', 'n'):
                    keeper_league_info(league_df, update.lower(), year)
                
                else:
                    print('Not a valid choice')
        
        else:
            print('Your league does not use keepers')
        
    while answer.lower() not in ('y', 'n'):
        answer = input('Would you like to run another report? [Y/N] ')
        if answer == 'y':
            return ''

        elif answer == 'n':
            return 'e'

        else:
            print('Wrong choice bub')

def basic_league_info(league_info):
    league_df = pd.DataFrame.from_dict(league_info)
    print('Creating the basic league info spreadsheet')
    league_df.loc['made the playoffs'] = np.where(league_df.loc['standing'] <= 6, 'Yes', 'No')
    league_df = league_df.drop(['division_id', 'streak_length', 'streak_type', 'logo_url', 'outcomes'])
    writer = pd.ExcelWriter(f'{write_path}temp_League_Results.xlsx')
    for t in league_df:
        team_name = league_df[t]['team_name']
        team_owner = league_df[t]['owner']

        #create basic team info table
        team_df = pd.DataFrame(columns=['Team Owner', 'Team Abbreviation', 'Wins', 'Losses', 'Total Points For', 
                                        'Total Points Against', 'Regular Season Finish', 'End of Season Rank',
                                        'Made the Playoffs'])
        team_df.set_index('Team Owner')
        team_dict = {
                    'Team Owner': league_df[t]['owner'],
                    'Team Abbreviation': league_df[t]['team_abbrev'],
                    'Wins': league_df[t]['wins'],
                    'Losses': league_df[t]['losses'],
                    'Total Points For': league_df[t]['points_for'],
                    'Total Points Against': league_df[t]['points_against'],
                    'Regular Season Finish': league_df[t]['standing'],
                    'End of Season Rank': league_df[t]['final_standing'],
                    'Made the Playoffs': league_df[t]['made the playoffs']
        }
        team_df = team_df.append(team_dict, ignore_index=True)

        #create Roster table
        players = league_df[t]['roster']
        players_df = pd.DataFrame(columns=['Name', 'Position', 'Pos. Rank'])
        for p in players:
            player_details = p.__dict__
            player_dict = {
                'Name': player_details['name'],
                'Position': player_details['position'],
                'Pos. Rank': player_details['posRank']
            }
            players_df = players_df.append(player_dict, ignore_index=True)

        #create schedule table
        schedule_df = pd.DataFrame(columns=['Week', 'Opponent', 'Team Score', 'Opponent\'s Score', 'Result'])
        w = 1
        for g in league_df[t]['schedule']:
            vs_details = g.__dict__
            if team_name == vs_details['team_name']:
                opponent = '**** 1st round BYE ****'
                outcome = 'BYE'
                team_score = ''
                opp_score = ''
            else:
                opponent = vs_details['team_name']
                team_score = league_df[t]['scores'][w - 1]
                opp_score = league_df[t]['scores'][w - 1] - league_df[t]['mov'][w - 1]
                if team_score > opp_score:
                    outcome = 'W'
                elif team_score == opp_score:
                    outcome = 'T'
                else:
                    outcome = 'L'

            schedule_dict = {
                'Week': w,
                'Opponent': opponent,
                'Team Score': team_score,
                'Opponent\'s Score': opp_score,
                'Result': outcome
            }

            w = w + 1
            
            schedule_df = schedule_df.append(schedule_dict, ignore_index=True)

        schedule_df.set_index('Week')

        sheetname = f'{team_name} ({team_owner})' 
        if len(sheetname) > 31:
            sheetname = team_name

        pattern = re.compile('[^A-Za-z0-9 _()]+')
        final_sheetname = pattern.sub('', sheetname)

        team_df.to_excel(writer, sheet_name=final_sheetname, startrow=1, index=False)
        players_df.to_excel(writer, sheet_name=final_sheetname, startrow=5, index=False)
        schedule_df.to_excel(writer, sheet_name=final_sheetname, startcol=4, startrow=5, index=False)

    writer.save()
    writer.handles = None

    workbook = load_workbook(filename = f'{write_path}temp_League_Results.xlsx')
    center_alignment = Alignment(horizontal='center')
    red_text = Font(color='00FF0000')
    red_dxf = DifferentialStyle(font=red_text, alignment=center_alignment)
    green_text = Font(color='00008000')
    green_dxf = DifferentialStyle(font=green_text, alignment=center_alignment)
    bye_text = Font(bold=True)
    bye_dxf = DifferentialStyle(font=bye_text, alignment=center_alignment)
    green_rule = Rule(type='containsText', operator='containsText', text='W', dxf=green_dxf)
    red_rule = Rule(type='containsText', operator='containsText', text='L', dxf=red_dxf)
    bye_rule = Rule(type='containsText', operator='containsText', text='BYE', dxf=bye_dxf)
    red_rule.formula = ['NOT(ISERROR(SEARCH("L",I7)))']
    green_rule.formula = ['NOT(ISERROR(SEARCH("W",I7)))']
    bye_rule.formula = ['NOT(ISERROR(SEARCH("BYE",I7)))']
    table_title_style = Font(bold=True, size=14)
    for sheets in workbook.sheetnames:
        ws = workbook[sheets]
        ws['E1'] = 'Team Info'
        ws['E1'].font = table_title_style
        ws['E1'].alignment = center_alignment
        ws['B5'] = 'Roster'
        ws['B5'].font = table_title_style
        ws['B5'].alignment = center_alignment
        ws['G5'] = 'Schedule'
        ws['G5'].font = table_title_style
        ws['G5'].alignment = center_alignment
        ws['I3'].alignment = center_alignment
        if ws['I3'].value == 'Yes':
            ws['I3'].font = Font(color='00008000', bold=True)
        elif ws['I3'].value == 'No':
            ws['I3'].font = Font(color='00FF0000', bold=True)

        ws.conditional_formatting.add('I7:I22', green_rule)
        ws.conditional_formatting.add('I7:I22', red_rule)
        ws.conditional_formatting.add('I7:I22', bye_rule)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2

    workbook.save(f'{write_path}{year}_League_Results.xlsx')
    os.remove(f'{write_path}temp_League_Results.xlsx')
    
def keeper_league_info(league_df, update, current_year):
    print('Creating the keeper spreadsheet')
    league_keeper_df = league_df.drop(['division_id', 'streak_length', 'streak_type', 'logo_url', 'team_abbrev', 
                                       'wins', 'losses', 'points_for', 'points_against', 'standing', 'final_standing', 
                                       'schedule', 'scores', 'outcomes', 'mov'])
    writer = pd.ExcelWriter(f'{write_path}tempLeague_Keeper_info.xlsx')
    for o in league_keeper_df:
        team_name = league_keeper_df[o]['team_name']
        team_owner = league_keeper_df[o]['owner']
        print(f'Creating player roster for {team_name}')
        players_df = pd.DataFrame(columns=['Name', 'Acquisition Type', 'Acquisition Value', 'Escalation Value'])
        print(f'Checking for current keepers for {team_name}')
        team_keepers = check_keepers(league_keeper_df[o]['team_id'], team_name, current_year)
        hist_year = current_year - 1

        for p in league_keeper_df[o].loc['roster']:
            player_details = p.__dict__
            player_id = player_details['playerId']
            player_name = player_details['name']
            print(f'Getting player information for {player_name}')
            player_json = espn_player_request(player_id, hist_year)
            transaction_number = 0
            player_acq = ''
            while player_acq not in ('DRAFT','WAIVER'):
                if player_acq == '':
                    player_acq = player_json['players'][0]['transactions'][transaction_number]['type']
            
                elif player_acq in ('TRADE_ACCEPT', 'ROSTER', 'FREEAGENT'):
                    transaction_number = transaction_number + 1
                    player_acq = player_json['players'][0]['transactions'][transaction_number]['type']

            player_value = player_json['players'][0]['transactions'][transaction_number]['bidAmount']
            player_esc_value = ceil(((player_value * .1) + player_value) + 5)
            keeper_status = 'No'
            if team_keepers:
                for keeper in team_keepers:
                    if player_id == keeper:
                        keeper_status = 'Yes'
                        print(f'+++++++ {player_name} has been selected as a keeper. +++++++')
                        break

            player_dict = {
                           'Name': player_name, 
                           'Acquisition Type': player_acq, 
                           'Acquisition Value': player_value, 
                           'Escalation Value': player_esc_value,
                           'Selected as Keeper': keeper_status
                          }
            players_df = players_df.append(player_dict, ignore_index=True)
            if update == 'y':
                headers = {'X-Fantasy-Source': 'kona',
                           'X-Fantasy-Filter': '{\"players\": {}}',
                           'Accept': 'application/json'
                          }
                cookies = {'SWID': f'{swid}', 'espn_s2': f'{espn_s2}', '_dcf': '1', 'region': 'unknown'}
                r = requests.post(url + f'{str(current_year)}/segments/0/leagues/{league_id}/players', headers=headers, 
                                  cookies=cookies, json=[{"id":player_id,"keeperValue":player_esc_value}])
                if r.status_code == 204:  # We want a 204 response
                    print(f'Successfully updated {player_name}\'s keeper value. Status code: {r.status_code}')

                else:
                    print(f'Something failed updating {player_name}\'s keeper value. Staus code: {r.status_code}')

        players_df.set_index('Name')
        sheetname = f'{team_name}({team_owner})'
        if len(sheetname) > 31:
            sheetname = f'{team_name}'
        
        try:
            players_df.to_excel(writer, sheet_name=f'{sheetname}', index=False) 
        except:
            sheetname = f'{team_owner}'
            players_df.to_excel(writer, sheet_name=f'{sheetname}', index=False)
        
    writer.save()
    writer.handles = None

    workbook = load_workbook(filename = f'{write_path}tempLeague_Keeper_info.xlsx')
    dims = {}
    for sheets in workbook.sheetnames:
        ws = workbook[sheets]
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2    

    workbook.save(f'{write_path}{current_year}_League_Keeper_info.xlsx')
    os.remove(f'{write_path}tempLeague_Keeper_info.xlsx')

def check_keepers(team_id, team_name, current_year):
    team_info = {}
    keepers = {}
    team_id_adj = team_id - 1
    year_str = str(current_year) 
    headers = {'X-Fantasy-Source': 'kona',
               'X-Fantasy-Filter': '{\"players\": {}}',
               'Accept': 'application/json'
               }
    cookies = {'SWID': f'{swid}', 'espn_s2': f'{espn_s2}', '_dcf': '1', 'region': 'unknown'}
    params = {'view':'mKeeperRosters'}
    r = requests.get(url + f'{year_str}/segments/0/leagues/{league_id}', params=params, headers=headers, cookies=cookies)
    keepers_json = r.json()
    team_info = keepers_json['teams'][team_id_adj]

    try:
        keepers = team_info['draftStrategy']['keeperPlayerIds']
    except KeyError:
        print('No keepers have been selected for ' + team_info['nickname'])
    
    if keepers:
        print(f'{team_name} has set {len(keepers)} Keepers.')

    return keepers    

def get_params():
    print('What would you like to do?, gather basic league info[L], gather keeper values [K], exit[E] ')
    answer = input('[L/K/E]: ')
    #print('[L/K/E]: ') # Used for debugging. Vscode debugger doesn't always handle inputs.
    current_year = 0
    if answer.lower() == 'e':
        return answer.lower(), 0
    
    while len(str(current_year)) != 4:
        try:
            current_year = int(input('What is the current league year, YYYY? '))
            if len(str(current_year)) == 4:
                break

            else:
                print('Please enter a valid year, YYYY')
                current_year = 0  

        except ValueError:
            print('Please enter a valid year, YYYY')
            continue
       
    return answer.lower(), current_year
    #return 'k' # Used for debugging. Vscode debugger doesn't always handle inputs.

if __name__ == "__main__":
    write_path, answer = init()
    while answer not in ('l', 'k', 'e', 'y') :
        answer, year = get_params()
        
        if answer in ('l', 'k'):
            answer = get_league_info(answer, year)

        elif answer == 'e':
            break
        
        else:
            print(f'Your answer was {answer}, Please choose either L, K, or E.')

    print('Thanks for using my app, please give a like and subscribe!')